#!/usr/bin/env python3
"""Build data.enc.json for the PRMT Client Security Status Dashboard.

Parses the monthly PromptQL/JumpCloud security-report workbook (Summary sheet +
one detail sheet per client) into JSON, then encrypts it AES-256-GCM with a
PBKDF2-SHA256 key (310k iterations) — the same "prmt-enc" format the dashboard
decrypts in-browser, so the public repo only ever holds ciphertext.

Usage:
    python3 scripts/build.py "<report.xlsx>" [--out data.enc.json]
                             [--plain data.json] [--password PASS]
                             [--duo duo.json] [--duo-map duo_map.json]

Password resolution order: --password, $PRMT_DASH_PASS, interactive prompt.
A fresh random salt is generated on every build (no encrypted side-files
depend on a stable key, unlike the Client Directory's images).

Duo merge: if duo.json exists (produced by scripts/duo_pull.py), each Duo
account is matched to a summary org by name (override in duo_map.json:
{"<duo account name>": "<org name>"}). Users are matched by email — an org's
"Active users WITHOUT MFA" list shrinks by everyone enrolled in Duo, MFA %
becomes the JumpCloud ∪ Duo union, and the posture score's MFA component is
re-weighted accordingly (same renormalized weights as the source report).
"""
import argparse, base64, getpass, json, os, re, sys, unicodedata
from datetime import date

import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

KDF_ITER = 310_000

SUMMARY_KEYS = [
    "name", "score", "totalUsers", "active", "suspended", "mfaEnrolled",
    "mfaPct", "activeNoMfa", "expiredPw", "adminMfa", "newUsers", "devices",
    "encrypted", "notEncrypted", "encUnknown", "notSeen7d", "notSeen30d",
    "mdm", "ssoApps", "unusedSso",
]
TABLE_HEADERS = {("Name", "Email", "Username"), ("SSO Application", "Users Assigned")}
SECTION_RE = re.compile(r"^[A-Z]\. ")


def slugify(s):
    s = unicodedata.normalize("NFKD", s)
    s = re.sub(r"[^\w\s-]", "", s).strip().lower()
    return re.sub(r"\s+", "-", s) or "org"


def cells(row):
    return [c for c in row if c is not None and str(c).strip() != ""]


def parse_pct(v):
    if v is None or v == "N/A":
        return None
    return float(str(v).rstrip("%"))


def parse_summary(ws):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    title = rows[0][0]
    meta_line = rows[1][0]
    methodology = rows[2][0]
    period = ""
    m = re.search(r"Reporting period:\s*([^|]+)", meta_line or "")
    if m:
        period = m.group(1).strip()
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Organization")
    orgs = []
    for r in rows[hdr_idx + 1:]:
        if not r or r[0] is None:
            continue
        rec = dict(zip(SUMMARY_KEYS, r))
        rec["score"] = parse_pct(rec["score"])
        rec["mfaPct"] = parse_pct(rec["mfaPct"])
        rec["slug"] = slugify(rec["name"])
        orgs.append(rec)
    return title, period, methodology, orgs


def parse_detail(ws):
    """Detail sheet -> {headline, sections:[{title, blocks:[kv|note|table]}]}."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    headline = ""
    if len(rows) > 2 and len(rows[2]) > 1 and rows[2][1]:
        headline = str(rows[2][1]).strip()
    sections, cur, table = [], None, None

    def close_table():
        nonlocal table
        table = None

    def ensure_section(title="General"):
        nonlocal cur
        if cur is None:
            cur = {"title": title, "blocks": []}
            sections.append(cur)
        return cur

    for r in rows[4:]:
        vals = cells(r)
        if not vals:
            close_table()
            continue
        a = str(vals[0]).strip()
        strs = tuple(str(v).strip() for v in vals)
        if SECTION_RE.match(a) and len(vals) == 1:
            close_table()
            cur = {"title": a, "blocks": []}
            sections.append(cur)
        elif strs in TABLE_HEADERS:
            table = {"type": "table", "title": None, "headers": list(strs), "rows": []}
            # attach a preceding standalone label (e.g. "Active users WITHOUT MFA (n)")
            blocks = ensure_section()["blocks"]
            if blocks and blocks[-1]["type"] == "label":
                table["title"] = blocks.pop()["text"]
            blocks.append(table)
        elif table is not None:
            table["rows"].append([str(v).strip() for v in strs])
        elif len(vals) >= 2:
            ensure_section()["blocks"].append(
                {"type": "kv", "k": a, "v": " — ".join(str(v).strip() for v in vals[1:])})
        elif a.startswith("Note:") or a.startswith("Trend"):
            ensure_section()["blocks"].append({"type": "note", "text": a})
        else:
            ensure_section()["blocks"].append({"type": "label", "text": a})
    return headline, sections


def build_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    title, period, methodology, orgs = parse_summary(wb["Summary"])
    by_name = {o["name"]: o for o in orgs}
    for sheet in wb.sheetnames[1:]:
        headline, sections = parse_detail(wb[sheet])
        org = by_name.get(sheet)
        if org is None:
            org = {"name": sheet, "slug": slugify(sheet)}
            orgs.append(org)
        org["headline"] = headline
        org["sections"] = sections
    return {
        "generatedAt": date.today().isoformat(),
        "title": title,
        "period": period,
        "source": "JumpCloud (live API via PromptQL)",
        "methodology": methodology,
        "caveat": ("Data source: JumpCloud only. Cisco Duo is not yet integrated — "
                   "MFA coverage for clients using Duo may be under-reported."),
        "orgs": orgs,
    }


def norm_name(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").casefold())


def mfa_weight(org):
    """Renormalized weight of the MFA component in the posture score.
    Mirrors the report's rule: unavailable metrics are excluded and the
    remaining weights renormalized (MFA 40, encryption 25, check-in 15,
    password 10, admin-MFA 10)."""
    if org.get("mfaPct") is None:
        return 0.0
    avail = [40]                                   # MFA
    if (org.get("devices") or 0) > 0:
        avail += [25, 15]                          # encryption + check-in
    if (org.get("active") or 0) > 0:
        avail += [10]                              # password compliance
    if org.get("adminMfa") in ("Yes", "No"):
        avail += [10]                              # admin MFA
    return 40 / sum(avail)


def merge_duo(data, duo, mapping):
    by_norm = {norm_name(o["name"]): o for o in data["orgs"]}
    merged = []
    for acct in duo["accounts"]:
        target = mapping.get(acct["name"], acct["name"])
        org = by_norm.get(norm_name(target))
        if org is None:
            print(f"  ! Duo account '{acct['name']}' matches no org — "
                  f"add it to duo_map.json", file=sys.stderr)
            continue
        enrolled = set(acct.get("enrolledEmails") or [])
        active = org.get("active") or 0

        # Shrink the JumpCloud "no MFA" list by everyone enrolled in Duo.
        covered = 0
        for sec in org.get("sections") or []:
            for blk in sec["blocks"]:
                if blk["type"] == "table" and "WITHOUT MFA" in (blk.get("title") or ""):
                    try:
                        ei = [h.lower() for h in blk["headers"]].index("email")
                    except ValueError:
                        continue
                    keep = [r for r in blk["rows"] if r[ei].lower() not in enrolled]
                    covered = len(blk["rows"]) - len(keep)
                    blk["rows"] = keep
                    blk["title"] = (f"Active users WITHOUT MFA — JumpCloud or Duo "
                                    f"({len(keep)})")

        old_pct = org.get("mfaPct")
        no_mfa = max(0, (org.get("activeNoMfa") or 0) - covered)
        new_pct = round((active - no_mfa) / active * 100, 1) if active else old_pct
        org["activeNoMfa"] = no_mfa
        org["mfaEnrolled"] = active - no_mfa
        org["mfaPct"] = new_pct
        if org.get("score") is not None and old_pct is not None and new_pct is not None:
            org["score"] = round(
                min(100, max(0, org["score"] + mfa_weight(org) * (new_pct - old_pct))), 1)

        if covered and new_pct is not None:
            org["headline"] = ((org.get("headline") or "").rstrip(". ") +
                               f" — {new_pct}% MFA coverage after including Duo").lstrip(" —")
        org["duo"] = {"account": acct["name"], "active": acct["active"],
                      "enrolled": acct["enrolled"], "mfaPct": acct["mfaPct"],
                      "coveredJcUsers": covered, "fetchedAt": duo["fetchedAt"]}
        for sec in org.get("sections") or []:
            if sec["title"].startswith("B."):
                sec["blocks"].append({"type": "kv", "k": "Duo account", "v": acct["name"]})
                sec["blocks"].append({"type": "kv", "k": "Duo MFA enrolled",
                                      "v": f"{acct['enrolled']} of {acct['active']} active"
                                           + (f" ({acct['mfaPct']}%)" if acct["mfaPct"] is not None else "")})
                sec["blocks"].append({"type": "kv", "k": "Combined MFA coverage (JumpCloud ∪ Duo)",
                                      "v": f"{active - no_mfa} of {active}"
                                           + (f" ({new_pct}%)" if new_pct is not None else "")})
                sec["blocks"].append({"type": "note", "text":
                    "Note: MFA coverage combines JumpCloud-native enrollment with Cisco Duo "
                    "enrollment, matched by email address. Posture score adjusted accordingly."})
                break
        merged.append(org["name"])

    if merged:
        data["source"] = "JumpCloud (live API via PromptQL) + Cisco Duo (Admin API)"
        data["caveat"] = (f"MFA coverage combines JumpCloud and Cisco Duo enrollment "
                          f"(Duo fetched {duo['fetchedAt']}: "
                          f"{', '.join(merged)}). Other orgs are JumpCloud-only.")
    return merged


def encrypt(password, obj):
    salt = os.urandom(16)
    iv = os.urandom(12)
    key = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt,
                     iterations=KDF_ITER).derive(password.encode())
    ct = AESGCM(key).encrypt(iv, json.dumps(obj, ensure_ascii=False).encode(), None)
    b64 = lambda b: base64.b64encode(b).decode()
    return {"format": "prmt-enc", "v": 1, "kdf": "PBKDF2-SHA256", "iter": KDF_ITER,
            "salt": b64(salt), "iv": b64(iv), "data": b64(ct)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", default="data.enc.json")
    ap.add_argument("--plain", default="data.json",
                    help="plaintext copy for local dev (git-ignored); '' to skip")
    ap.add_argument("--password", default=None)
    ap.add_argument("--duo", default="duo.json",
                    help="duo.json from scripts/duo_pull.py; merged if the file exists")
    ap.add_argument("--duo-map", default="duo_map.json",
                    help="optional {duo account name: org name} overrides")
    args = ap.parse_args()

    pw = args.password or os.environ.get("PRMT_DASH_PASS") or getpass.getpass("Dashboard password: ")
    if not pw:
        sys.exit("A password is required.")
    data = build_data(args.xlsx)
    if args.duo and os.path.exists(args.duo):
        mapping = json.load(open(args.duo_map)) if os.path.exists(args.duo_map) else {}
        merged = merge_duo(data, json.load(open(args.duo)), mapping)
        print(f"merged Duo MFA into {len(merged)} org(s): {', '.join(merged) or '—'}")
    if args.plain:
        with open(args.plain, "w") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
        print(f"wrote {args.plain} (plaintext — git-ignored, never commit)")
    with open(args.out, "w") as f:
        json.dump(encrypt(pw, data), f, indent=1)
    print(f"wrote {args.out}  ({len(data['orgs'])} orgs, period: {data['period']})")


if __name__ == "__main__":
    main()
